"""
Excel Parser

Parses .xlsx files with support for:
- Multi-sheet detection (each sheet = week/phase)
- Header row auto-detection
- 1RM extraction from header blocks
- Day boundary detection ("Day 1", "Day 2" patterns)
- Formula evaluation for calculated weights
"""

import io
import re
import logging
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from .base import BaseParser
from .models import (
    ParseResult,
    ParsedWorkout,
    ParsedExercise,
    ColumnInfo,
    FileInfo,
    ExerciseFlag,
)

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parser for Excel (.xlsx) files"""

    # Common header patterns
    EXERCISE_HEADERS = ['exercise', 'movement', 'lift', 'name']
    SETS_HEADERS = ['sets', 'set']
    REPS_HEADERS = ['reps', 'rep', 'repetitions']
    WEIGHT_HEADERS = ['weight', 'load', 'kg', 'lbs', 'pounds']
    REST_HEADERS = ['rest', 'pause', 'recovery']
    NOTES_HEADERS = ['notes', 'note', 'comments', 'instructions']

    # 1RM detection patterns
    ONE_RM_PATTERN = re.compile(r'1\s*RM|one\s*rep\s*max|max', re.IGNORECASE)
    EXERCISE_1RM_PATTERN = re.compile(
        r'([A-Za-z\s]+?)[\s:=]+(\d+(?:\.\d+)?)\s*(kg|lbs?)?',
        re.IGNORECASE
    )

    def can_parse(self, file_info: FileInfo) -> bool:
        """Check if this parser can handle the file"""
        return file_info.extension.lower() in ['.xlsx', '.xls']

    async def parse(self, content: bytes, file_info: FileInfo) -> ParseResult:
        """Parse Excel file and return normalized workout data"""
        self.errors = []
        self.warnings = []

        try:
            # Load workbook
            wb = load_workbook(io.BytesIO(content), data_only=True)

            result = ParseResult(
                sheet_names=wb.sheetnames,
                detected_format="excel_multi_sheet" if len(wb.sheetnames) > 1 else "excel_single_sheet"
            )

            all_workouts = []
            total_rows = 0

            # Process each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                total_rows += ws.max_row

                # Try to extract 1RMs from sheet header
                one_rms = self._extract_one_rms(ws)
                if one_rms:
                    result.one_rep_maxes.update(one_rms)

                # Detect header row and columns
                header_row, columns = self._detect_header_row(ws)
                if header_row is None:
                    self.add_warning(f"Could not detect header row in sheet '{sheet_name}'")
                    continue

                if not result.header_row:
                    result.header_row = header_row

                result.columns.extend(columns)

                # Parse workouts from this sheet
                workouts = self._parse_sheet(ws, sheet_name, header_row, columns)
                all_workouts.extend(workouts)

            result.workouts = all_workouts
            result.total_rows = total_rows

            # Detect patterns across all workouts
            result.patterns = self.detect_patterns(all_workouts)

            # Calculate confidence
            result.confidence = self._calculate_confidence(result)

            result.errors = self.errors
            result.warnings = self.warnings
            result.success = len(self.errors) == 0

            return result

        except Exception as e:
            logger.exception(f"Failed to parse Excel file: {e}")
            return ParseResult(
                success=False,
                errors=[f"Failed to parse Excel file: {str(e)}"],
                confidence=0
            )

    def _extract_one_rms(self, ws: Worksheet) -> Dict[str, float]:
        """Extract 1RM values from sheet header area (first 10 rows)"""
        one_rms = {}

        for row_idx in range(1, min(11, ws.max_row + 1)):
            for col_idx in range(1, min(10, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = str(cell.value or "").strip()

                # Look for 1RM patterns
                if self.ONE_RM_PATTERN.search(cell_value):
                    # Check adjacent cells for exercise: weight pairs
                    for adj_row in range(row_idx, min(row_idx + 5, ws.max_row + 1)):
                        row_text = " ".join(
                            str(ws.cell(row=adj_row, column=c).value or "")
                            for c in range(1, ws.max_column + 1)
                        )
                        matches = self.EXERCISE_1RM_PATTERN.findall(row_text)
                        for match in matches:
                            exercise = match[0].strip()
                            weight = float(match[1])
                            if exercise and weight > 0:
                                one_rms[exercise] = weight

        return one_rms

    def _detect_header_row(self, ws: Worksheet) -> Tuple[Optional[int], List[ColumnInfo]]:
        """Detect the header row and map columns"""
        columns = []
        best_row = None
        best_score = 0

        # Search first 20 rows for header
        for row_idx in range(1, min(21, ws.max_row + 1)):
            row_values = [
                str(ws.cell(row=row_idx, column=col).value or "").lower().strip()
                for col in range(1, ws.max_column + 1)
            ]

            score = 0
            row_columns = []

            for col_idx, value in enumerate(row_values, 1):
                col_info = ColumnInfo(
                    index=col_idx,
                    name=str(ws.cell(row=row_idx, column=col_idx).value or f"Column {col_idx}"),
                    sample_values=[]
                )

                # Check for known header patterns
                if any(h in value for h in self.EXERCISE_HEADERS):
                    col_info.detected_type = "exercise"
                    col_info.confidence = 90
                    score += 3
                elif any(h in value for h in self.SETS_HEADERS):
                    col_info.detected_type = "sets"
                    col_info.confidence = 85
                    score += 2
                elif any(h in value for h in self.REPS_HEADERS):
                    col_info.detected_type = "reps"
                    col_info.confidence = 85
                    score += 2
                elif any(h in value for h in self.WEIGHT_HEADERS):
                    col_info.detected_type = "weight"
                    col_info.confidence = 85
                    score += 2
                elif any(h in value for h in self.REST_HEADERS):
                    col_info.detected_type = "rest"
                    col_info.confidence = 80
                    score += 1
                elif any(h in value for h in self.NOTES_HEADERS):
                    col_info.detected_type = "notes"
                    col_info.confidence = 80
                    score += 1

                # Get sample values from next few rows
                for sample_row in range(row_idx + 1, min(row_idx + 4, ws.max_row + 1)):
                    sample_val = ws.cell(row=sample_row, column=col_idx).value
                    if sample_val is not None:
                        col_info.sample_values.append(str(sample_val)[:50])

                row_columns.append(col_info)

            if score > best_score:
                best_score = score
                best_row = row_idx
                columns = row_columns

        return best_row, columns

    def _parse_sheet(
        self,
        ws: Worksheet,
        sheet_name: str,
        header_row: int,
        columns: List[ColumnInfo]
    ) -> List[ParsedWorkout]:
        """Parse workouts from a single sheet"""
        workouts = []
        current_workout = None
        current_day = None
        exercise_order = 1

        # Find column indices
        exercise_col = next((c.index for c in columns if c.detected_type == "exercise"), None)
        sets_col = next((c.index for c in columns if c.detected_type == "sets"), None)
        reps_col = next((c.index for c in columns if c.detected_type == "reps"), None)
        weight_col = next((c.index for c in columns if c.detected_type == "weight"), None)
        rest_col = next((c.index for c in columns if c.detected_type == "rest"), None)
        notes_col = next((c.index for c in columns if c.detected_type == "notes"), None)

        if not exercise_col:
            self.add_warning(f"No exercise column found in sheet '{sheet_name}'")
            return workouts

        # Parse data rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            # Check for day boundary
            first_cell = str(ws.cell(row=row_idx, column=1).value or "").strip()
            day_match = self.DAY_PATTERN.search(first_cell)
            week_match = self.WEEK_PATTERN.search(first_cell)

            if day_match or week_match:
                # Save current workout and start new one
                if current_workout and current_workout.exercises:
                    workouts.append(current_workout)

                day_num = day_match.group(1) if day_match else None
                week_num = week_match.group(1) if week_match else None

                current_workout = ParsedWorkout(
                    name=f"{sheet_name} - {first_cell}",
                    week=f"Week {week_num}" if week_num else None,
                    day=f"Day {day_num}" if day_num else first_cell,
                    source_sheet=sheet_name,
                    source_row_start=row_idx,
                )
                exercise_order = 1
                continue

            # Get exercise name
            exercise_name = ws.cell(row=row_idx, column=exercise_col).value
            if not exercise_name or str(exercise_name).strip() == "":
                continue

            exercise_name = self.normalize_exercise_name(str(exercise_name))

            # Create workout if not exists
            if not current_workout:
                current_workout = ParsedWorkout(
                    name=sheet_name,
                    source_sheet=sheet_name,
                    source_row_start=row_idx,
                )

            # Parse exercise data
            sets_val = ws.cell(row=row_idx, column=sets_col).value if sets_col else 1
            reps_val = ws.cell(row=row_idx, column=reps_col).value if reps_col else "1"
            weight_val = ws.cell(row=row_idx, column=weight_col).value if weight_col else None
            rest_val = ws.cell(row=row_idx, column=rest_col).value if rest_col else None
            notes_val = ws.cell(row=row_idx, column=notes_col).value if notes_col else None

            # Parse reps and detect flags
            reps_str, reps_flags = self.parse_reps(str(reps_val or "1"))

            # Parse weight and detect flags
            weight_str, weight_unit, weight_flags = self.parse_weight(str(weight_val) if weight_val else "")

            # Check for RPE in notes
            rpe = None
            if notes_val:
                rpe = self.parse_rpe(str(notes_val))

            # Determine order (check for superset notation)
            order = str(exercise_order)
            # Check if exercise name starts with letter notation like "A.", "B.", "5a."
            order_match = re.match(r'^([0-9]+[a-z]?|[A-Z])[\.\):\s]', exercise_name)
            if order_match:
                order = order_match.group(1)
                exercise_name = exercise_name[len(order_match.group(0)):].strip()

            # Combine flags
            flags = list(set(reps_flags + weight_flags))

            # Check for warmup indicators
            if notes_val and re.search(r'warm\s*up|wu|warmup', str(notes_val), re.IGNORECASE):
                flags.append(ExerciseFlag.WARMUP)

            exercise = ParsedExercise(
                raw_name=exercise_name,
                order=order,
                sets=int(sets_val) if sets_val and str(sets_val).isdigit() else 1,
                reps=reps_str,
                weight=weight_str,
                weight_unit=weight_unit,
                rpe=rpe,
                notes=str(notes_val) if notes_val else None,
                rest_seconds=int(rest_val) if rest_val and str(rest_val).isdigit() else None,
                flags=flags,
            )

            current_workout.exercises.append(exercise)
            current_workout.source_row_end = row_idx
            exercise_order += 1

        # Add last workout
        if current_workout and current_workout.exercises:
            # Detect superset groups
            current_workout.exercises = self.detect_superset_groups(current_workout.exercises)
            workouts.append(current_workout)

        return workouts

    def _calculate_confidence(self, result: ParseResult) -> float:
        """Calculate overall parsing confidence"""
        if not result.workouts:
            return 0

        total_exercises = sum(len(w.exercises) for w in result.workouts)
        if total_exercises == 0:
            return 0

        # Base confidence
        confidence = 50

        # Bonus for detecting header row
        if result.header_row:
            confidence += 10

        # Bonus for finding exercise column
        if any(c.detected_type == "exercise" for c in result.columns):
            confidence += 15

        # Bonus for finding other columns
        if any(c.detected_type == "sets" for c in result.columns):
            confidence += 5
        if any(c.detected_type == "reps" for c in result.columns):
            confidence += 5
        if any(c.detected_type == "weight" for c in result.columns):
            confidence += 5

        # Bonus for 1RMs
        if result.one_rep_maxes:
            confidence += 5

        # Bonus for multi-sheet with structure
        if len(result.sheet_names) > 1 and len(result.workouts) > 1:
            confidence += 5

        return min(confidence, 100)
